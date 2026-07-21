"""Service-level tests for the main V1 workflows."""

from __future__ import annotations

import sqlite3
import tempfile
import unittest
from contextlib import closing
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from database import create_database_backup
from exercise_management import create_exercise, list_exercises, update_exercise
from export_management import build_history_workbook
from init_db import initialize_database
from profile_management import ProfileError, create_profile, list_profiles, update_profile
from progress_calculations import (
    days_since_previous_workout,
    exercise_progress,
    history_dataframe,
    latest_exercise_progress,
    workout_comparison_dataframe,
)
from routine_management import (
    add_exercise_to_workout,
    create_routine,
    create_workout,
    duplicate_routine,
    list_routines,
    list_workout_exercises,
    list_workouts,
    move_workout_exercise,
    remove_exercise_from_workout,
    update_routine,
    update_workout_exercise,
)
from workout_logging import (
    delete_completed_session,
    delete_logged_set,
    ExerciseLog,
    SetEntry,
    WorkoutLogError,
    get_previous_result,
    get_previous_results,
    get_session_review,
    list_sessions,
    save_completed_session,
    update_logged_set,
)


class WorkflowTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_path = Path(self.temp_dir.name) / "test.db"
        initialize_database(self.db_path, seed_sample_routine=False)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_intensity_reps_cannot_exceed_twenty(self) -> None:
        with self.assertRaisesRegex(WorkoutLogError, "cannot exceed 20"):
            update_logged_set(
                -1,
                weight=50,
                reps=8,
                intensity_method="Rest-pause",
                intensity_reps=21,
                db_path=self.db_path,
            )

    def test_exercise_create_edit_and_deactivate(self) -> None:
        exercise_id = create_exercise(
            "Custom Press", "Shoulders", "Dumbbells", db_path=self.db_path
        )
        update_exercise(
            exercise_id,
            name="Custom Overhead Press",
            primary_muscle_group="Shoulders",
            equipment="Dumbbells",
            notes="Controlled reps",
            active=False,
            db_path=self.db_path,
        )
        exercise = next(item for item in list_exercises(db_path=self.db_path) if item.id == exercise_id)
        self.assertEqual(exercise.name, "Custom Overhead Press")
        self.assertFalse(exercise.active)

    def test_routine_build_log_history_progress_and_snapshots(self) -> None:
        exercises = list_exercises(active_only=True, db_path=self.db_path)
        squat = next(item for item in exercises if item.name == "Barbell Back Squat")
        bench = next(item for item in exercises if item.name == "Bench Press")
        routine_id = create_routine("Strength", db_path=self.db_path)
        workout_id = create_workout(routine_id, "Day One", db_path=self.db_path)
        squat_config = add_exercise_to_workout(
            workout_id,
            squat.id,
            target_min_reps=5,
            target_max_reps=8,
            target_sets=3,
            db_path=self.db_path,
        )
        bench_config = add_exercise_to_workout(
            workout_id,
            bench.id,
            target_min_reps=6,
            target_max_reps=10,
            target_sets=3,
            db_path=self.db_path,
        )
        move_workout_exercise(bench_config, -1, self.db_path)
        configured = list_workout_exercises(workout_id, self.db_path)
        self.assertEqual([item.exercise_name for item in configured], ["Bench Press", "Barbell Back Squat"])

        session_day = date.today() - timedelta(days=3)
        session_id = save_completed_session(
            routine_id,
            workout_id,
            session_day,
            datetime.combine(session_day, datetime.min.time()).replace(hour=18),
            [
                ExerciseLog(
                    bench_config,
                    (SetEntry(60, 8), SetEntry(60, 7, "Rest-pause", 2)),
                ),
                ExerciseLog(squat_config, (SetEntry(80, 6),)),
            ],
            notes="Good session",
            db_path=self.db_path,
        )

        update_routine(
            routine_id,
            name="Renamed Strength",
            description="Edited later",
            active=True,
            db_path=self.db_path,
        )
        update_workout_exercise(
            bench_config,
            target_min_reps=10,
            target_max_reps=12,
            target_sets=4,
            instructions="New target",
            db_path=self.db_path,
        )
        remove_exercise_from_workout(squat_config, self.db_path)

        review = get_session_review(session_id, self.db_path)
        self.assertEqual(review["routine_name"], "Strength")
        self.assertEqual(review["exercises"][0]["target_min_reps"], 6)
        self.assertEqual(len(review["exercises"]), 2)
        self.assertEqual(len(list_sessions(exercise_id=bench.id, db_path=self.db_path)), 1)
        previous = get_previous_result(bench.id, db_path=self.db_path)
        self.assertIsNotNone(previous)
        self.assertEqual(len(previous["sets"]), 2)
        self.assertEqual(previous["sets"][1]["intensity_reps"], 2)
        previous_results = get_previous_results(
            [bench.id, squat.id], db_path=self.db_path
        )
        self.assertEqual(set(previous_results), {bench.id, squat.id})
        self.assertEqual(previous_results[squat.id]["sets"][0]["reps"], 6)
        latest = latest_exercise_progress(db_path=self.db_path)
        self.assertEqual(set(latest["exercise"]), {"Bench Press", "Barbell Back Squat"})
        comparison = workout_comparison_dataframe(
            routine_id=routine_id,
            workout_ids=[workout_id],
            db_path=self.db_path,
        )
        self.assertEqual(len(comparison), 2)
        self.assertEqual(
            set(comparison["exercise"]),
            {"Bench Press", "Barbell Back Squat"},
        )
        progress = exercise_progress(bench.id, db_path=self.db_path)
        self.assertEqual(
            list(progress.columns),
            ["date", "workout_name", "weight", "reps", "evaluation"],
        )
        self.assertEqual(progress.iloc[0]["weight"], 60.0)
        self.assertEqual(progress.iloc[0]["reps"], 8)
        self.assertEqual(days_since_previous_workout(db_path=self.db_path), 3)

    def test_schema_has_expected_tables(self) -> None:
        with closing(sqlite3.connect(self.db_path)) as connection:
            tables = {
                row[0]
                for row in connection.execute(
                    "SELECT name FROM sqlite_master WHERE type = 'table'"
                )
            }
        expected = {
            "profiles",
            "exercises",
            "routines",
            "workouts",
            "workout_exercises",
            "workout_sessions",
            "session_exercises",
            "logged_sets",
            "schema_migrations",
        }
        self.assertTrue(expected.issubset(tables))

    def test_profiles_keep_history_and_progress_separate(self) -> None:
        jozef_id = create_profile("Jozef", db_path=self.db_path)
        adrian_id = create_profile("Adrian", db_path=self.db_path)
        with self.assertRaises(ProfileError):
            create_profile("jozef", db_path=self.db_path)
        bench = next(
            item
            for item in list_exercises(active_only=True, db_path=self.db_path)
            if item.name == "Bench Press"
        )
        routine_id = create_routine("Shared Plan", db_path=self.db_path)
        workout_id = create_workout(routine_id, "Bench Day", db_path=self.db_path)
        config_id = add_exercise_to_workout(
            workout_id,
            bench.id,
            target_min_reps=5,
            target_max_reps=10,
            target_sets=1,
            db_path=self.db_path,
        )
        workout_date = date(2026, 2, 1)
        for profile_id, weight, reps in ((jozef_id, 50, 8), (adrian_id, 70, 5)):
            save_completed_session(
                routine_id,
                workout_id,
                workout_date,
                datetime.combine(workout_date, datetime.min.time()).replace(hour=18),
                [ExerciseLog(config_id, (SetEntry(weight, reps),))],
                profile_id=profile_id,
                db_path=self.db_path,
            )

        self.assertEqual(len(list_sessions(profile_id=jozef_id, db_path=self.db_path)), 1)
        self.assertEqual(len(list_sessions(profile_id=adrian_id, db_path=self.db_path)), 1)
        self.assertEqual(
            get_previous_result(
                bench.id, profile_id=jozef_id, db_path=self.db_path
            )["sets"][0]["weight"],
            50,
        )
        self.assertEqual(
            get_previous_result(
                bench.id, profile_id=adrian_id, db_path=self.db_path
            )["sets"][0]["weight"],
            70,
        )
        self.assertEqual(
            exercise_progress(
                bench.id, profile_id=jozef_id, db_path=self.db_path
            ).iloc[0]["weight"],
            50,
        )
        adrian_history = history_dataframe(
            profile_id=adrian_id, db_path=self.db_path
        )
        self.assertEqual(adrian_history["profile"].unique().tolist(), ["Adrian"])

        update_profile(
            adrian_id,
            name="Adrian",
            notes="Archived client",
            active=False,
            db_path=self.db_path,
        )
        active_names = {
            profile.name for profile in list_profiles(active_only=True, db_path=self.db_path)
        }
        self.assertIn("Jozef", active_names)
        self.assertNotIn("Adrian", active_names)
        self.assertEqual(len(list_sessions(profile_id=adrian_id, db_path=self.db_path)), 1)

    def test_progress_evaluates_weight_and_reps_together(self) -> None:
        bench = next(
            item
            for item in list_exercises(active_only=True, db_path=self.db_path)
            if item.name == "Bench Press"
        )
        routine_id = create_routine("Progress Test", db_path=self.db_path)
        workout_id = create_workout(routine_id, "Bench Day", db_path=self.db_path)
        config_id = add_exercise_to_workout(
            workout_id,
            bench.id,
            target_min_reps=6,
            target_max_reps=10,
            target_sets=1,
            db_path=self.db_path,
        )
        results = [(50, 10), (50, 10), (50, 9), (52.5, 8), (50, 9)]
        for offset, (weight, reps) in enumerate(results):
            workout_date = date(2026, 1, 1) + timedelta(days=offset)
            save_completed_session(
                routine_id,
                workout_id,
                workout_date,
                datetime.combine(workout_date, datetime.min.time()).replace(hour=18),
                [ExerciseLog(config_id, (SetEntry(weight, reps),))],
                db_path=self.db_path,
            )

        progress = exercise_progress(bench.id, db_path=self.db_path)
        self.assertEqual(
            progress["evaluation"].tolist(),
            ["Baseline", "No progress", "Regression", "Progress", "Progress"],
        )

    def test_one_set_defaults_and_routine_duplication(self) -> None:
        initialize_database(self.db_path, seed_sample_routine=True)
        sample = next(
            routine
            for routine in list_routines(db_path=self.db_path)
            if routine.name == "Sample Full Body"
        )
        sample_workout = list_workouts(sample.id, db_path=self.db_path)[0]
        sample_exercises = list_workout_exercises(sample_workout.id, self.db_path)
        self.assertTrue(sample_exercises)
        self.assertTrue(all(item.target_sets == 1 for item in sample_exercises))

        with sqlite3.connect(self.db_path) as connection:
            connection.execute(
                "UPDATE workout_exercises SET target_sets = 4 WHERE id = ?",
                (sample_exercises[0].id,),
            )
            connection.execute(
                "DELETE FROM schema_migrations WHERE version = ?",
                ("v5_heavy_duty_fixed_one_set",),
            )
            connection.commit()
        initialize_database(self.db_path, seed_sample_routine=False)
        migrated = list_workout_exercises(sample_workout.id, self.db_path)
        self.assertTrue(all(item.target_sets == 1 for item in migrated))

        copy_id = duplicate_routine(sample.id, "My Full Body", db_path=self.db_path)
        copied_workout = list_workouts(copy_id, db_path=self.db_path)[0]
        copied_exercises = list_workout_exercises(copied_workout.id, self.db_path)
        self.assertEqual(
            [item.exercise_name for item in copied_exercises],
            [item.exercise_name for item in sample_exercises],
        )
        self.assertEqual(list_sessions(routine_id=copy_id, db_path=self.db_path), [])

    def test_correct_delete_and_export_logged_set(self) -> None:
        bench = next(
            item
            for item in list_exercises(active_only=True, db_path=self.db_path)
            if item.name == "Bench Press"
        )
        routine_id = create_routine("Export Test", db_path=self.db_path)
        workout_id = create_workout(routine_id, "Bench Day", db_path=self.db_path)
        config_id = add_exercise_to_workout(
            workout_id,
            bench.id,
            target_min_reps=6,
            target_max_reps=10,
            target_sets=1,
            db_path=self.db_path,
        )
        session_id = save_completed_session(
            routine_id,
            workout_id,
            date.today(),
            datetime.now(),
            [ExerciseLog(config_id, (SetEntry(50, 8), SetEntry(50, 7)))],
            db_path=self.db_path,
        )
        review = get_session_review(session_id, self.db_path)
        first_set, second_set = review["exercises"][0]["sets"]
        with self.assertRaisesRegex(
            WorkoutLogError, "Weight must use 0.25 kg increments"
        ):
            update_logged_set(
                first_set["id"],
                weight=55.6,
                reps=9,
                db_path=self.db_path,
            )
        update_logged_set(
            first_set["id"],
            weight=55.75,
            reps=9,
            intensity_method="Forced",
            intensity_reps=2,
            notes="Corrected",
            db_path=self.db_path,
        )
        delete_logged_set(second_set["id"], db_path=self.db_path)

        corrected = get_session_review(session_id, self.db_path)["exercises"][0]["sets"]
        self.assertEqual(len(corrected), 1)
        self.assertEqual(corrected[0]["weight"], 55.75)
        self.assertEqual(corrected[0]["reps"], 9)
        self.assertEqual(corrected[0]["intensity_method"], "Forced")
        self.assertEqual(corrected[0]["intensity_reps"], 2)
        exported = history_dataframe(
            routine_id=routine_id, exercise_id=bench.id, db_path=self.db_path
        )
        self.assertEqual(len(exported), 1)
        self.assertEqual(exported.iloc[0]["notes"], "Corrected")
        self.assertEqual(exported.iloc[0]["intensity_reps"], 2)
        self.assertNotIn("session_id", exported.columns)

        sessions = list_sessions(routine_id=routine_id, db_path=self.db_path)
        workbook_bytes = build_history_workbook(sessions, exported)
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Sessions", "Sets"])
        self.assertNotIn("Session ID", [cell.value for cell in workbook["Sessions"][1]])
        set_headers = [cell.value for cell in workbook["Sets"][1]]
        self.assertNotIn("Session ID", set_headers)
        self.assertIn("Intensity Reps", set_headers)
        self.assertEqual(workbook["Sessions"]["D2"].value, "Export Test")
        self.assertEqual(workbook["Sets"]["G2"].value, 55.75)
        self.assertEqual(workbook["Sets"]["G2"].number_format, "0.##")
        self.assertEqual(workbook["Sets"].freeze_panes, "A2")

        backup = create_database_backup(self.db_path)
        backup_path = Path(self.temp_dir.name) / "restored.db"
        backup_path.write_bytes(backup)
        with closing(sqlite3.connect(backup_path)) as connection:
            session_count = connection.execute(
                "SELECT COUNT(*) FROM workout_sessions"
            ).fetchone()[0]
        self.assertEqual(session_count, 1)

        with self.assertRaisesRegex(WorkoutLogError, "Completed workout not found"):
            delete_completed_session(
                session_id,
                profile_id=999,
                db_path=self.db_path,
            )
        self.assertEqual(
            len(list_sessions(routine_id=routine_id, db_path=self.db_path)),
            1,
        )
        delete_completed_session(session_id, db_path=self.db_path)
        self.assertEqual(
            list_sessions(routine_id=routine_id, db_path=self.db_path),
            [],
        )
        with sqlite3.connect(self.db_path) as connection:
            session_exercise_count = connection.execute(
                "SELECT COUNT(*) FROM session_exercises"
            ).fetchone()[0]
            logged_set_count = connection.execute(
                "SELECT COUNT(*) FROM logged_sets"
            ).fetchone()[0]
            workout_count = connection.execute(
                "SELECT COUNT(*) FROM workouts WHERE id = ?", (workout_id,)
            ).fetchone()[0]
        self.assertEqual(session_exercise_count, 0)
        self.assertEqual(logged_set_count, 0)
        self.assertEqual(workout_count, 1)
        with self.assertRaisesRegex(WorkoutLogError, "Completed workout not found"):
            delete_completed_session(session_id, db_path=self.db_path)


if __name__ == "__main__":
    unittest.main()
