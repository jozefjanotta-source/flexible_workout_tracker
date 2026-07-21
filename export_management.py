"""Formatted Excel export helpers for workout history."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Iterable, Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_BORDER = Border(bottom=Side(style="thin", color="A6A6A6"))


def build_history_workbook(
    sessions: Iterable[Mapping[str, object]], sets: pd.DataFrame
) -> bytes:
    """Create a two-sheet, Excel-compatible workout-history workbook."""
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Sessions"
    sets_sheet = workbook.create_sheet("Sets")

    session_columns = [
        "workout_date",
        "completed_at",
        "profile_name",
        "routine_name",
        "workout_name",
        "exercise_count",
        "set_count",
        "notes",
    ]
    session_headers = [
        "Workout Date",
        "Completed At",
        "Profile",
        "Routine",
        "Workout",
        "Exercises",
        "Sets",
        "Session Notes",
    ]
    session_rows = []
    for row in sessions:
        values = [_excel_value(row.get(column)) for column in session_columns]
        values[0] = _as_date(row.get("workout_date"))
        values[1] = _as_datetime(row.get("completed_at"))
        session_rows.append(values)
    _write_sheet(summary_sheet, session_headers, session_rows, "SessionsTable")

    set_headers = [
        "Date",
        "Profile",
        "Routine",
        "Workout",
        "Exercise",
        "Set",
        "Weight",
        "Reps",
        "Intensity Method",
        "Intensity Reps",
        "Set Notes",
    ]
    set_columns = [
        "date",
        "profile",
        "routine",
        "workout",
        "exercise",
        "set_number",
        "weight",
        "reps",
        "intensity_method",
        "intensity_reps",
        "notes",
    ]
    set_rows = []
    for row in sets.to_dict("records"):
        values = [_excel_value(row.get(column)) for column in set_columns]
        values[0] = _as_date(row.get("date"))
        set_rows.append(values)
    _write_sheet(sets_sheet, set_headers, set_rows, "SetsTable")
    for cell in sets_sheet["G"][1:]:
        cell.number_format = "0.0#"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _excel_value(value: object) -> object:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "item"):
        return value.item()  # Convert NumPy scalar values.
    return value


def _as_date(value: object) -> object:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value[:10])
        except ValueError:
            return value
    return _excel_value(value)


def _as_datetime(value: object) -> object:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return value
    return _excel_value(value)


def _write_sheet(
    sheet: object,
    headers: list[str],
    rows: list[list[object]],
    table_name: str,
) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="left")

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(sheet.max_row, 1)}"

    if rows:
        table = Table(
            displayName=table_name,
            ref=f"A1:{get_column_letter(len(headers))}{sheet.max_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = "dd/mm/yy hh:mm"
            elif isinstance(cell.value, date):
                cell.number_format = "dd/mm/yy"

    for index, header in enumerate(headers, start=1):
        values = [header] + [str(row[index - 1] or "") for row in rows]
        width = min(max(len(value) for value in values) + 2, 42)
        sheet.column_dimensions[get_column_letter(index)].width = max(width, 10)
